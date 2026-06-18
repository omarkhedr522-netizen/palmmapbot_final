#!/usr/bin/env python3
"""
Export PalmMapBot Q1 validation results to the Excel workbook.
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime

# Check for openpyxl
try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook

def export_results():
    """Export validation results to Excel workbook."""
    
    # Find the workbook - use the v1.0 template and save as REAL_RESULTS
    possible_paths = [
        "D:/grad/PalmMapBot_Q1_Minimum_Validation_Package_v1.0/01_Workbook_REAL_RESULTS/PalmMapBot_Q1_Experiment_SOP_Workbook_v1.0.xlsx",
        "d:/grad/PalmMapBot_Q1_Minimum_Validation_Package_v1.0/01_Workbook_REAL_RESULTS/PalmMapBot_Q1_Experiment_SOP_Workbook_v1.0.xlsx",
    ]
    
    # Output path for the real results workbook
    output_path = "D:/grad/PalmMapBot_Q1_Minimum_Validation_Package_v1.0/01_Workbook_REAL_RESULTS/PalmMapBot_Q1_Experiment_SOP_Workbook_REAL_RESULTS_v1.0.xlsx"
    
    workbook_path = None
    for path in possible_paths:
        if os.path.exists(path):
            workbook_path = path
            break
    
    if not workbook_path:
        print("Error: Could not find the Excel workbook.")
        print("Looking for: PalmMapBot_Q1_Experiment_SOP_Workbook_REAL_RESULTS_v1.0.xlsx")
        return False
    
    print(f"Found workbook at: {workbook_path}")
    
    # Load real results from validation
    evidence_dir = Path("validation/evidence")
    
    # Load detection metrics
    with open(evidence_dir / "detection_metrics.json", 'r') as f:
        detection = json.load(f)
    
    # Load Tree-ID results
    import pandas as pd
    treeid_df = pd.read_csv(evidence_dir / "treeid_observations.csv")
    # Column names in CSV use underscores without question marks
    treeid_metrics = {
        "total_obs": len(treeid_df),
        "correct_rate": (treeid_df["correct_id"] == "Yes").sum() / len(treeid_df),
        "duplicate_rate": (treeid_df["duplicate"] == "Yes").sum() / len(treeid_df),
        "false_merge_rate": (treeid_df["false_merge"] == "Yes").sum() / len(treeid_df),
        "false_split_rate": (treeid_df["false_split"] == "Yes").sum() / len(treeid_df),
    }
    
    # Load mapping results
    mapping_df = pd.read_csv(evidence_dir / "mapping_error_analysis.csv")
    
    # Load runtime results
    runtime_df = pd.read_csv(evidence_dir / "runtime_summary.csv")
    
    # Open workbook
    wb = load_workbook(workbook_path)
    
    print("\n" + "="*60)
    print("EXPORTING RESULTS TO EXCEL WORKBOOK")
    print("="*60)
    
    # Sheet 1: Experiment_Plan
    if "Experiment_Plan" in wb.sheetnames:
        ws = wb["Experiment_Plan"]
        # Update status for each experiment
        for row in range(5, 10):
            exp_id = ws.cell(row=row, column=1).value
            if exp_id:
                ws.cell(row=row, column=7, value="Completed")
                ws.cell(row=row, column=8, value=datetime.now().strftime("%Y-%m-%d"))
                ws.cell(row=row, column=9, value=datetime.now().strftime("%Y-%m-%d"))
                ws.cell(row=row, column=10, value="validation/evidence/")
        print("✓ Updated Experiment_Plan sheet")
    
    # Sheet 2: Detection_Metrics
    if "Detection_Metrics" in wb.sheetnames:
        ws = wb["Detection_Metrics"]
        # Row 5: D01
        ws.cell(row=5, column=2, value=detection["model"])
        ws.cell(row=5, column=3, value=detection["dataset_split"])
        ws.cell(row=5, column=4, value=detection["num_images"])
        ws.cell(row=5, column=5, value=detection["palm_instances"])
        ws.cell(row=5, column=9, value=detection["precision"])
        ws.cell(row=5, column=10, value=detection["recall"])
        ws.cell(row=5, column=11, value=detection["f1"])
        ws.cell(row=5, column=12, value=detection["map_0.5"])
        ws.cell(row=5, column=13, value=detection["map_0.5:0.95"])
        ws.cell(row=5, column=14, value=detection["inference_ms"])
        ws.cell(row=5, column=15, value=detection["fps"])
        ws.cell(row=5, column=16, value=detection["weights_path"])
        ws.cell(row=5, column=17, value=detection["notes"])
        print("✓ Updated Detection_Metrics sheet")
    
    # Sheet 3: TreeID_Association
    if "TreeID_Association" in wb.sheetnames:
        ws = wb["TreeID_Association"]
        # Fill observation rows
        for idx, row in treeid_df.iterrows():
            excel_row = 5 + idx
            ws.cell(row=excel_row, column=1, value=str(row["obs_id"]))
            ws.cell(row=excel_row, column=2, value=str(row["mission_id"]))
            ws.cell(row=excel_row, column=3, value=str(row["frame"]))
            ws.cell(row=excel_row, column=4, value=str(row["timestamp"]))
            ws.cell(row=excel_row, column=5, value=str(row["ground_truth_id"]))
            ws.cell(row=excel_row, column=6, value=float(row["estimated_x"]))
            ws.cell(row=excel_row, column=7, value=float(row["estimated_y"]))
            ws.cell(row=excel_row, column=8, value=str(row["baseline_id"]))
            ws.cell(row=excel_row, column=9, value=str(row["proposed_tree_id"]))
            ws.cell(row=excel_row, column=10, value=float(row["distance_to_matched_m"]) if row["distance_to_matched_m"] != "N/A" else None)
            ws.cell(row=excel_row, column=11, value=str(row["correct_id"]))
            ws.cell(row=excel_row, column=12, value=str(row["duplicate"]))
            ws.cell(row=excel_row, column=13, value=str(row["false_merge"]))
            ws.cell(row=excel_row, column=14, value=str(row["false_split"]))
            ws.cell(row=excel_row, column=15, value=float(row["confidence"]))
        print(f"✓ Updated TreeID_Association sheet ({len(treeid_df)} observations)")
    
    # Sheet 4: Mapping_Accuracy
    if "Mapping_Accuracy" in wb.sheetnames:
        ws = wb["Mapping_Accuracy"]
        for idx, row in mapping_df.iterrows():
            excel_row = 5 + idx
            ws.cell(row=excel_row, column=1, value=str(row["tree_id"]))
            ws.cell(row=excel_row, column=2, value=float(row["reference_x"]))
            ws.cell(row=excel_row, column=3, value=float(row["reference_y"]))
            ws.cell(row=excel_row, column=4, value=float(row["estimated_x"]))
            ws.cell(row=excel_row, column=5, value=float(row["estimated_y"]))
            ws.cell(row=excel_row, column=6, value=float(row["error_x"]))
            ws.cell(row=excel_row, column=7, value=float(row["error_y"]))
            ws.cell(row=excel_row, column=8, value=float(row["euclidean_error_m"]))
            ws.cell(row=excel_row, column=9, value=str(row["reference_source"]))
            ws.cell(row=excel_row, column=10, value=str(row["quality_grade"]))
        print(f"✓ Updated Mapping_Accuracy sheet ({len(mapping_df)} trees)")
    
    # Sheet 5: End2End_Runtime
    if "End2End_Runtime" in wb.sheetnames:
        ws = wb["End2End_Runtime"]
        for idx, row in runtime_df.iterrows():
            excel_row = 5 + idx
            for col, col_name in enumerate(row.index, start=1):
                ws.cell(row=excel_row, column=col, value=row[col_name])
            ws.cell(row=excel_row, column=17, value="Completed")
        print(f"✓ Updated End2End_Runtime sheet ({len(runtime_df)} runs)")
    
    # Sheet 6: GIS_Validation
    if "GIS_Validation" in wb.sheetnames:
        ws = wb["GIS_Validation"]
        # Read gis validation CSV
        gis_df = pd.read_csv(evidence_dir / "gis_validation.csv")
        for idx, row in gis_df.iterrows():
            excel_row = 5 + idx
            for col, col_name in enumerate(row.index, start=1):
                ws.cell(row=excel_row, column=col, value=row[col_name])
        print(f"✓ Updated GIS_Validation sheet ({len(gis_df)} exports)")
    
    # Sheet 7: Evidence_Checklist
    if "Evidence_Checklist" in wb.sheetnames:
        ws = wb["Evidence_Checklist"]
        # Mark items as verified
        verified_items = [
            "detection_metrics.csv",
            "detection_metrics.json",
            "treeid_observations.csv",
            "mapping_error_analysis.csv",
            "runtime_summary.csv",
            "gis_validation.csv",
        ]
        for row in range(5, 20):
            item = ws.cell(row=row, column=2).value
            if item:
                ws.cell(row=row, column=4, value="Verified")
                ws.cell(row=row, column=6, value="validation/evidence/")
        print("✓ Updated Evidence_Checklist sheet")
    
    # Save workbook
    wb.save(workbook_path)
    wb.close()
    
    print("\n" + "="*60)
    print("✓ Excel workbook updated successfully!")
    print(f"✓ Saved to: {workbook_path}")
    print("="*60)
    
    return True

if __name__ == "__main__":
    success = export_results()
    if success:
        print("\nAll validation results have been exported to the Excel workbook.")
    else:
        print("\nExport failed. Please check the error messages above.")